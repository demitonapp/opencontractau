"""
City of Sydney awarded contract register scraper.

Source:    cityofsydney.nsw.gov.au GIPA contracts register
Format:    XLSX download (openpyxl), discovered from the GIPA page
Threshold: AU$150,000 (NSW GIPA Act 2009, schedule 2, clause 4)
ABN:       Not disclosed -- supplier name only
Updates:   Quarterly

Best-in-class NSW council GIPA register: XLSX format with consistent
column structure. The direct download URL contains a date suffix
(e.g. gipa-register-2026-04-10.xlsx) so it is discovered at runtime
by scanning the GIPA page for .xlsx hrefs.

Verified columns (April 2026):
  Date | Description | Name of Supplier | Value of Contract ($) |
  Contract Number | Contract Period
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime

import httpx

from opencontractsau.models.ocds import Publisher, Release, ReleasePackage
from opencontractsau.scrapers.base import BROWSER_UA
from opencontractsau.transformers.council import (
    CouncilContractRow,
    parse_au_date,
    parse_value,
    row_to_release,
)

logger = logging.getLogger(__name__)

GIPA_PAGE_URL = (
    "https://www.cityofsydney.nsw.gov.au/council-governance-administration"
    "/contracts-over-150000-awarded-by-city-of-sydney"
)
BASE_URL = "https://www.cityofsydney.nsw.gov.au"

COUNCIL_KEY = "SYDNEY_COUNCIL"
COUNCIL_NAME = "City of Sydney"

_XLSX_PATTERN = re.compile(r'href="([^"]*gipa[^"]*\.xlsx)"', re.IGNORECASE)
_LINK_PATTERN = re.compile(r'href="([^"]*contracts[^"]*150[^"]*\.xlsx|[^"]*gipa[^"]*register[^"]*\.xlsx)"', re.IGNORECASE)

_COL_DATE = ("date", "award date", "date awarded", "contract date")
_COL_TITLE = ("description", "goods or services", "title", "subject")
_COL_SUPPLIER = ("supplier", "name of supplier", "contractor", "awarded to", "vendor")
_COL_VALUE = ("value", "contract value", "amount", "$")
_COL_REF = ("contract number", "reference", "contract ref", "contract no", "number")


def _match_col(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        if any(c in hl for c in candidates):
            return i
    return None


def _parse_xlsx(xlsx_bytes: bytes) -> list[CouncilContractRow]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("SYDNEY_COUNCIL: openpyxl not installed")
        return []

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows: list[tuple] = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Find the header row: first row where at least 3 cells are non-empty strings
    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for c in row if c and str(c).strip())
        if non_empty >= 3:
            header_row_idx = i
            break

    headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]

    col_date = _match_col(headers, _COL_DATE)
    col_title = _match_col(headers, _COL_TITLE)
    col_supplier = _match_col(headers, _COL_SUPPLIER)
    col_value = _match_col(headers, _COL_VALUE)
    col_ref = _match_col(headers, _COL_REF)

    if col_supplier is None:
        logger.warning("SYDNEY_COUNCIL: could not find supplier column. Headers: %s", headers)
        return []

    results: list[CouncilContractRow] = []
    for data_row in rows[header_row_idx + 1:]:
        if not any(c for c in data_row if c):
            continue

        def _cell(idx: int | None) -> str:
            if idx is None or idx >= len(data_row):
                return ""
            v = data_row[idx]
            return str(v).strip() if v is not None else ""

        supplier = _cell(col_supplier)
        if not supplier:
            continue

        title = _cell(col_title)
        value_raw = _cell(col_value)
        date_raw = _cell(col_date)
        ref = _cell(col_ref) or None

        # Dates may come as datetime objects from Excel
        award_date: datetime | None = None
        d = data_row[col_date] if col_date is not None and col_date < len(data_row) else None
        if isinstance(d, datetime):
            award_date = d
        elif date_raw:
            award_date = parse_au_date(date_raw)

        results.append(CouncilContractRow(
            council_key=COUNCIL_KEY,
            council_name=COUNCIL_NAME,
            reference=ref,
            title=title or f"City of Sydney Contract - {supplier}",
            awarded_to=supplier,
            value_aud=parse_value(value_raw),
            award_date=award_date,
        ))

    logger.info("SYDNEY_COUNCIL: parsed %d rows", len(results))
    return results


async def scrape(**kwargs) -> ReleasePackage:
    """Discover and parse the City of Sydney GIPA contracts register (XLSX)."""
    async with httpx.AsyncClient(
        timeout=60.0,
        headers={"User-Agent": BROWSER_UA},
        follow_redirects=True,
    ) as client:
        # Step 1: discover the XLSX URL from the GIPA page
        xlsx_url: str | None = None
        try:
            page_resp = await client.get(GIPA_PAGE_URL)
            page_resp.raise_for_status()
            m = _XLSX_PATTERN.search(page_resp.text)
            if not m:
                m = _LINK_PATTERN.search(page_resp.text)
            if m:
                href = m.group(1)
                xlsx_url = href if href.startswith("http") else BASE_URL + href
        except Exception as exc:
            logger.warning("SYDNEY_COUNCIL: GIPA page discovery failed: %s", exc)

        if not xlsx_url:
            logger.error("SYDNEY_COUNCIL: could not discover XLSX download URL")
            return ReleasePackage(
                uri=f"https://github.com/demitonapp/opencontractsau/releases/{COUNCIL_KEY}",
                publishedDate=datetime.utcnow(),
                publisher=Publisher(),
                releases=[],
            )

        # Step 2: download XLSX
        logger.info("SYDNEY_COUNCIL: downloading %s", xlsx_url)
        try:
            xlsx_resp = await client.get(xlsx_url)
            xlsx_resp.raise_for_status()
            xlsx_bytes = xlsx_resp.content
        except Exception as exc:
            logger.error("SYDNEY_COUNCIL: XLSX download failed: %s", exc)
            return ReleasePackage(
                uri=f"https://github.com/demitonapp/opencontractsau/releases/{COUNCIL_KEY}",
                publishedDate=datetime.utcnow(),
                publisher=Publisher(),
                releases=[],
            )

    rows = _parse_xlsx(xlsx_bytes)
    releases: list[Release] = [r for seq, row in enumerate(rows, 1) if (r := row_to_release(row, seq=seq))]

    logger.info("SYDNEY_COUNCIL: %d releases ready", len(releases))
    return ReleasePackage(
        uri=f"https://github.com/demitonapp/opencontractsau/releases/{COUNCIL_KEY}",
        publishedDate=datetime.utcnow(),
        publisher=Publisher(),
        releases=releases,
    )
