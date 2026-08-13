"""City of Sydney GIPA XLSX register parsing.

Regression coverage for the two real bugs found live (2026-08-13):
wb.active pointing at the wrong sheet, and "Successful Tenderer Name" /
"Contract Name" headers not matching the original column-name tuples.
"""

import io

import openpyxl

from opencontractau.scrapers.nsw.councils.city_of_sydney import _XLSX_PATTERN, _parse_xlsx


def _workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "1 - Cover"
    cover.append(["City of Sydney", "Contract Register"])

    data = wb.create_sheet("2 - All GIPA Contracts")
    data.append(["Part 1: Register of all contracts valued at or above $150,000"])
    data.append([
        "Contract\nClass", "Quote or Tender\nReference", "Contract ID",
        "Successful Tenderer Name", "Contract Name",
        "Effective date of\ncontract", "Estimated amount payable under the contract\n(Contract Value)",
    ])
    data.append(["Class 1", "RFT 1539", "CT001932", "Acme Cloud Pty Ltd", "Cloud Hosting Services",
                 __import__("datetime").datetime(2016, 2, 16), 2176784])

    wb.active = wb.sheetnames.index("1 - Cover")  # the file's "last saved" tab, not the data sheet
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSheetSelection:
    def test_picks_the_all_gipa_contracts_sheet_not_wb_active(self):
        rows = _parse_xlsx(_workbook_bytes())
        assert len(rows) == 1
        assert rows[0].awarded_to == "Acme Cloud Pty Ltd"


class TestHeaderMatching:
    def test_tenderer_and_contract_name_headers_are_recognised(self):
        rows = _parse_xlsx(_workbook_bytes())
        row = rows[0]
        assert row.title == "Cloud Hosting Services"
        # "Quote or Tender Reference" matches the ref-column candidates before
        # "Contract ID" does (leftmost header wins) - same column the live
        # register picks, verified against the real August 2026 file.
        assert row.reference == "RFT 1539"
        assert row.value_aud == 2176784


class TestXlsxDiscoveryRegex:
    def test_matches_a_link_with_a_download_query_string(self):
        html = (
            '<a href="https://www.cityofsydney.nsw.gov.au/.../gipa-report-august.xlsx'
            '?download=true">download</a>'
        )
        m = _XLSX_PATTERN.search(html)
        assert m is not None
        assert m.group(1).endswith("gipa-report-august.xlsx")
